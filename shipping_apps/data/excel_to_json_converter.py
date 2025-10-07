"""
سكريبت تحويل aramex.xlsx إلى aramex.json
استخدمه مرة واحدة فقط لتحويل البيانات
"""

from openpyxl import load_workbook
import json
import os

def convert_excel_to_json():
    """تحويل ملف Excel إلى JSON"""
    
    print("🔄 بدء التحويل...")
    
    # قراءة ملف Excel
    try:
        wb = load_workbook("aramex.xlsx", data_only=True)
        sheet = wb.active
        print(f"✅ تم فتح ملف Excel")
    except FileNotFoundError:
        print("❌ خطأ: ملف aramex.xlsx غير موجود")
        return False
    except Exception as e:
        print(f"❌ خطأ في قراءة Excel: {e}")
        return False
    
    # تحويل البيانات
    data = []
    rows_count = 0
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        country_en = row[0]
        country_ar = row[1]
        first_half = row[2]
        
        # تخطي الصفوف الفارغة
        if not (country_en or country_ar) or first_half is None:
            continue
        
        data.append({
            "Country_EN": str(country_en) if country_en else "",
            "Country_AR": str(country_ar) if country_ar else "",
            "FirstHalfKg": float(row[2]) if row[2] is not None else 0,
            "Add_0_5_to_10": float(row[3]) if row[3] is not None else 0,
            "Add_10_to_15": float(row[4]) if row[4] is not None else 0,
            "Add_over_15": float(row[5]) if row[5] is not None else 0
        })
        rows_count += 1
    
    print(f"✅ تم تحويل {rows_count} دولة")
    
    # حفظ في JSON
    try:
        # حفظ في المجلد الحالي
        with open('aramex.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"✅ تم حفظ aramex.json في المجلد الحالي")
        
        # حفظ نسخة في مجلد data
        if not os.path.exists('data'):
            os.makedirs('data')
        with open('data/aramex.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"✅ تم حفظ نسخة في data/aramex.json")
        
        # عرض إحصائيات
        print(f"\n📊 إحصائيات:")
        print(f"   - عدد الدول: {len(data)}")
        print(f"   - حجم الملف: {os.path.getsize('aramex.json') / 1024:.2f} KB")
        
        # عرض أول 3 دول
        print(f"\n📝 أول 3 دول:")
        for i, country in enumerate(data[:3], 1):
            print(f"   {i}. {country['Country_AR']} ({country['Country_EN']})")
            print(f"      السعر الأساسي: {country['FirstHalfKg']} ريال")
        
        return True
        
    except Exception as e:
        print(f"❌ خطأ في حفظ JSON: {e}")
        return False

if __name__ == "__main__":
    print("=" * 50)
    print("  📦 تحويل aramex.xlsx إلى aramex.json")
    print("=" * 50)
    print()
    
    success = convert_excel_to_json()
    
    print()
    print("=" * 50)
    if success:
        print("✅ تم التحويل بنجاح!")
        print()
        print("📋 الخطوات التالية:")
        print("   1. احذف ملف aramex.xlsx (لم تعد بحاجة له)")
        print("   2. استخدم aramex.json في التطبيق")
        print("   3. عدّل main.py لقراءة JSON بدلاً من Excel")
    else:
        print("❌ فشل التحويل!")
    print("=" * 50)
